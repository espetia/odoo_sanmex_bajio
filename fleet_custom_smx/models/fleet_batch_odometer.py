# -*- coding: utf-8 -*-
import base64
import io
from datetime import datetime
import logging

from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError

try:
    import openpyxl
except ImportError:
    openpyxl = None

_logger = logging.getLogger(__name__)


class FleetBatchOdometer(models.Model):
    _name = "fleet.batch.odometer"
    _inherit = ["mail.thread", "mail.activity.mixin"]
    _description = "Fleet Batch Odometer"

    name = fields.Char("Nombre", readonly=True, required=True, default=lambda self: _("Nuevo"))
    date_to = fields.Date("Fecha inicio", required=True)
    date_from = fields.Date("Fecha final", required=True)
    company_id = fields.Many2one("res.company", string="Company", required=True, default=lambda self: self.env.company)
    currency_id = fields.Many2one("res.currency", string="Currency", related="company_id.currency_id", readonly=True)
    user_id = fields.Many2one("res.users", string="Usuario", required=True, default=lambda self: self.env.user)
    state = fields.Selection(
        [("draft", "Borrador"), ("confirmed", "Confirmado"), ("cancelled", "Cancelado")],
        string="Estado",
        default="draft",
        tracking=True,
    )
    line_ids = fields.One2many("fleet.batch.odometer.line", "batch_id", string="Lineas de odometros")
    excel_file = fields.Binary("Archivo Excel", attachment=True, help="Subir un archivo de excel .xlsx")
    excel_filename = fields.Char("Nombre del archivo")
    import_log = fields.Text("Log de importación", readonly=True)
    total_lines = fields.Integer("Total líneas", compute="_compute_total_lines")
    odometers_ids = fields.One2many(
        "fleet.vehicle.odometer", "batch_id", string="Vehicle Odometers", help="Odometros creados con este batch"
    )
    odometers_count = fields.Integer(
        "Total Odometros", compute="_compute_odometers_count", help="Numero de odometros relacionados al batch"
    )

    @api.model
    def create(self, vals):
        if vals.get("name", "Nuevo") == "Nuevo":
            vals["name"] = self.env["ir.sequence"].next_by_code("batch.upload.odometer") or "Nuevo"
        return super().create(vals)

    @api.onchange("date_to", "date_from")
    def _onchange_date_to_date_from(self):
        """Validar que date_to no sea mayor que date_from"""
        if self.date_to and self.date_from and self.date_to > self.date_from:
            raise ValidationError(_("La fecha inicio no puede ser mayor que la fecha final."))

    def unlink(self):
        """Prevenir eliminación si el batch está confirmado"""
        for record in self:
            if record.state in ["confirmed"]:
                raise UserError(
                    "No se puede eliminar el batch, porque está en estado Confirmado. Solo se pueden eliminar batches en estado 'Borrador' o 'Cancelado'."
                )
        return super().unlink()

    @api.depends("odometers_ids")
    def _compute_odometers_count(self):
        for record in self:
            record.odometers_count = len(record.odometers_ids)

    @api.depends("line_ids")
    def _compute_total_lines(self):
        for record in self:
            record.total_lines = len(record.line_ids)

    def action_confirm(self):
        for record in self:
            if record.state != "draft":
                raise UserError(_("Solo los registros en borrador pueden ser confirmados."))
            if not record.line_ids:
                raise UserError(_("No hay lineas para procesar. Por favor agregar lineas antes de confirmar."))

            # Validar que el odómetro sea mayor al último registrado
            for line in record.line_ids:
                # Buscar el último odómetro registrado del vehículo
                last_odometer = self.env["fleet.vehicle.odometer"].search(
                    [("vehicle_id", "=", line.vehicle_id.id), ("date", "<=", line.date or fields.Date.today())],
                    order="date desc, id desc",
                    limit=1,
                )

                # Validar que el nuevo valor sea mayor
                if last_odometer and line.odometer_value <= last_odometer.value:
                    raise ValidationError(
                        _("Odómetro (%s) debe ser más grande que el registro anterior (%s) para el vehículo %s.")
                        % (line.odometer_value, last_odometer.value, line.vehicle_id.name)
                    )

                # Validación adicional: odómetro no puede ser negativo o cero
                if line.odometer_value <= 0:
                    raise ValidationError(
                        _("El odómetro debe ser mayor a 0 para el vehículo %s.") % line.vehicle_id.name
                    )

            # Crear registros de odómetro en fleet.vehicle.odometer
            record._create_vehicle_odometers()

            record.state = "confirmed"
            record.message_post(body=_("Batch Odometer confirmed with %d lines.") % len(record.line_ids))
            # _logger.info("Batch Odometer %s confirmed.", record.name)

    def action_cancel(self):
        for record in self:
            if record.state != "draft":
                raise UserError(_("Only draft records can be cancelled."))
            record.state = "cancelled"

            # *** ELIMINAR ODÓMETROS CREADOS ANTES DE CANCELAR ***
            if record.odometers_ids:
                try:
                    record.odometers_ids.unlink()
                    record.message_post(
                        body=_("Se eliminaron %d registros de odómetro al cancelar el lote.")
                        % len(record.odometers_ids)
                    )
                    _logger.info(
                        "Eliminados %d odómetros del lote cancelado %s", len(record.odometers_ids), record.name
                    )
                except Exception as e:
                    _logger.error("Error eliminando odómetros del lote %s: %s", record.name, str(e))
                    # Continuar con la cancelación aunque falle la eliminación

                # Limpiar la relación
                record.odometers_ids = [(5, 0, 0)]
            # _logger.info("Batch Odometer %s cancelled.", record.name)

    def action_reset_to_draft(self):
        for record in self:
            if record.state not in ["confirmed", "cancelled"]:
                raise UserError(_("Only confirmed or cancelled records can be reset to draft."))

            # *** SI ESTÁ CONFIRMADO, ELIMINAR ODÓMETROS ANTES DE RESETEAR ***
            if record.state == "confirmed" and record.odometers_ids:
                try:
                    odometer_count = len(record.odometers_ids)
                    record.odometers_ids.unlink()
                    record.message_post(
                        body=_("Se eliminaron %d registros de odómetro al resetear a borrador.") % odometer_count
                    )
                    _logger.info("Eliminados %d odómetros al resetear lote %s", odometer_count, record.name)
                except Exception as e:
                    _logger.error("Error eliminando odómetros del lote %s: %s", record.name, str(e))

                # Limpiar la relación
                record.odometers_ids = [(5, 0, 0)]
            record.state = "draft"
            record.message_post(body=_("Batch Odometer reset to draft."))
            # _logger.info("Batch Odometer %s reset to draft.", record.name)

    def action_read_excel(self):
        """Leer y mostrar las líneas del Excel sin procesarlas"""
        if not self.excel_file:
            raise UserError(_("Please upload an Excel file first."))

        if not self.excel_filename.lower().endswith((".xlsx", ".xls")):
            raise UserError(_("Only .xlsx and .xls files are supported."))

        if not openpyxl:
            raise UserError(_("Please install openpyxl library to read Excel files: pip install openpyxl"))

        try:
            # Decodificar el archivo
            file_content = base64.b64decode(self.excel_file)

            # Leer el archivo Excel con configuración flexible
            excel_data = self._read_xlsx_file(file_content)

            # Mostrar información del archivo
            log_lines = []
            log_lines.append("=== INFORMACIÓN DEL ARCHIVO EXCEL ===")
            log_lines.append(f"Archivo: {self.excel_filename}")
            log_lines.append(f"Fila de encabezados: {excel_data['header_row']}")
            log_lines.append(f"Última fila de datos: {excel_data['data_end_row']}")
            if excel_data.get("footer_excluded", 0) > 0:
                log_lines.append(f"Filas de footer excluidas: {excel_data['footer_excluded']}")
            log_lines.append(f"Total filas de datos: {len(excel_data['rows'])}")
            log_lines.append(f"Total columnas: {len(excel_data['headers'])}")
            log_lines.append("")

            log_lines.append("=== ENCABEZADOS DETECTADOS ===")
            for i, header in enumerate(excel_data["headers"]):
                log_lines.append(f"Columna {i+1}: '{header}'")
            log_lines.append("")

            # Detectar formato de Excel
            excel_format = self._detect_excel_format(excel_data["headers"])
            log_lines.append(f"=== FORMATO DETECTADO: {excel_format.upper()} ===")
            log_lines.append("")

            log_lines.append("=== PRIMERAS 10 FILAS DE DATOS ===")
            for row_num, row_data in enumerate(excel_data["rows"][:10], 1):
                log_lines.append(f"Fila {row_num + excel_data['header_row']}:")
                for i, value in enumerate(row_data):
                    header = excel_data["headers"][i] if i < len(excel_data["headers"]) else f"Col_{i+1}"
                    log_lines.append(f"  {header}: '{value}'")
                log_lines.append("")

            if len(excel_data["rows"]) > 10:
                log_lines.append(f"... y {len(excel_data['rows']) - 10} filas más")

            self.import_log = "\n".join(log_lines)

            return {
                "type": "ir.actions.client",
                "tag": "display_notification",
                "params": {
                    "title": _("Excel Leído Correctamente"),
                    "message": _("Se han leído %d filas del archivo Excel. Formato: %s")
                    % (len(excel_data["rows"]), excel_format.upper()),
                    "type": "success",
                    "sticky": False,
                },
            }

        except Exception as e:
            error_msg = _("Error reading Excel file: %s") % str(e)
            self.import_log = error_msg
            _logger.error("Excel read error: %s", str(e))
            raise UserError(error_msg)

    def action_import_excel(self):
        """Importar datos desde archivo Excel (.xlsx)"""
        if not self.excel_file:
            raise UserError(_("Por favor subir un excel."))

        if not self.excel_filename.lower().endswith(".xlsx"):
            raise UserError(_("Solo archivos .xlsx soportados."))

        if not openpyxl:
            raise UserError(_("Please install openpyxl library to import Excel files: pip install openpyxl"))

        try:
            # Decodificar el archivo
            file_content = base64.b64decode(self.excel_file)

            # Procesar archivo .xlsx
            lines_data = self._process_xlsx_file(file_content)
            # Limpiar líneas existentes
            self.line_ids.unlink()

            # Crear nuevas líneas
            success_count = 0
            error_log = []

            for row_num, line_data in enumerate(lines_data, start=1):  # Empezar desde fila 2 (después del header)
                try:
                    self._create_odometer_line(line_data)
                    success_count += 1
                except Exception as e:
                    error_msg = _("Fila %d: %s") % (row_num, str(e))
                    error_log.append(error_msg)
                    _logger.warning("Error importando fila %d: %s", row_num, str(e))

            # Actualizar log de importación
            log_msg = _("Importación completada.\nCorrecto: %d lineas\nErrores: %d") % (success_count, len(error_log))
            if error_log:
                log_msg += "\n\nErrores:\n" + "\n".join(error_log)

            self.import_log = log_msg

            if success_count > 0:
                self.message_post(body=_("Excel importado: %d lines importadas correctamente.") % success_count)

            return {
                "type": "ir.actions.client",
                "tag": "display_notification",
                "params": {
                    "title": _("Importación Completada"),
                    "message": _("Se importaron %d líneas correctamente.") % success_count,
                    "type": "success",
                    "sticky": False,
                },
            }

        except Exception as e:
            error_msg = _("Error processing Excel file: %s") % str(e)
            self.import_log = error_msg
            _logger.error("Excel import error: %s", str(e))
            raise UserError(error_msg)

    def _detect_header_row(self, sheet, max_rows=10):
        """Detectar en qué fila están los encabezados"""
        for row_num in range(1, min(max_rows + 1, sheet.max_row + 1)):
            row = list(sheet.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]

            # Buscar patrones comunes de encabezados
            header_indicators = [
                "nota_id",
                "placas",
                "placa",
                "fecha",
                "odometro",
                "cantidad",
                "monto",
                "descripcion",
                "producto",
                "rendimiento",
                "vehiculo",
                "vehículo",
                "odómetro",
                "no economico",
                "tipo mov",
                "id_cliente",
            ]

            if row and any(cell for cell in row):
                # Filtrar valores nulos/vacíos para columnas fusionadas
                row_text = [str(cell).lower().strip() for cell in row if cell is not None and str(cell).strip()]
                matches = sum(1 for indicator in header_indicators for cell in row_text if indicator in cell)

                # Para el formato de Cargas_odometros.xlsx - buscar combinación específica
                has_nota_id = any("nota_id" in cell for cell in row_text)
                has_placas = any("placas" in cell for cell in row_text)
                has_odometro = any("odometro" in cell or "odómetro" in cell for cell in row_text)
                has_tipo_mov = any("tipo mov" in cell for cell in row_text)

                # Si encuentra la combinación específica del archivo de cargas
                if has_nota_id and has_placas and has_odometro:
                    return row_num

                if has_tipo_mov and has_nota_id:
                    return row_num

                # Para Petromayab - buscar combinación específica
                has_vehiculo = any("vehículo" in cell or "vehiculo" in cell for cell in row_text)
                has_no_economico = any("no economico" in cell for cell in row_text)

                if has_vehiculo and has_placas and has_odometro:
                    return row_num

                # Si encuentra al menos 4 coincidencias, probablemente sea la fila de encabezados
                if matches >= 4:
                    return row_num

        # Si no encuentra patrones, asumir que la primera fila es el encabezado
        return 1

    def _detect_footer_start(self, sheet, header_row):
        """Detectar dónde empieza el footer para no procesarlo"""
        # Patrones comunes de footer
        footer_indicators = [
            "total",
            "suma",
            "subtotal",
            "grand total",
            "resumen",
            "summary",
            "totales",
            "sumatoria",
            "consolidado",
            "fin",
            "end",
            "final",
            "página",
            "page",
            "fecha de reporte",
            "report date",
            "generado",
            "promedio",
            "average",
            "máximo",
            "mínimo",
            "count",
            "contador",
            # Patrones específicos para diferentes formatos
            "de 1",
            "1 de 1",
            "página 1",
            "página 1 de 1",  # Petromayab
        ]

        data_end_row = sheet.max_row
        consecutive_empty = 0
        last_data_row = header_row

        for row_num in range(header_row + 1, sheet.max_row + 1):
            row = list(sheet.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]

            # Si la fila está completamente vacía
            if not any(cell for cell in row if cell is not None and str(cell).strip()):
                consecutive_empty += 1
                # Si hay 2 o más filas vacías consecutivas, probablemente terminaron los datos
                if consecutive_empty >= 2:
                    data_end_row = row_num - consecutive_empty
                    break
            else:
                consecutive_empty = 0

                # Verificar si es una fila de footer
                row_text = [str(cell).lower().strip() if cell else "" for cell in row]
                is_footer = False

                # Buscar indicadores de footer
                for indicator in footer_indicators:
                    for cell_text in row_text:
                        if indicator in cell_text:
                            is_footer = True
                            break
                    if is_footer:
                        break

                # Verificar patrones específicos de Petromayab: "Página X de Y"
                if not is_footer:
                    first_cell = str(row[0]).strip() if row[0] else ""
                    if "página" in first_cell.lower() and "de" in first_cell.lower():
                        is_footer = True

                # Verificar patrones específicos de fecha/hora en footer (formato: DD/MM/YYYY HH:MM:SS)
                if not is_footer:
                    first_cell = str(row[0]).strip() if row[0] else ""
                    # Detectar timestamps como "03/09/2025 22:04:52"
                    if "/" in first_cell and ":" in first_cell and len(first_cell) > 10:
                        is_footer = True

                # Verificar patrones de footer (como filas con solo totales numéricos)
                if not is_footer:
                    first_cell = str(row[0]).lower().strip() if row[0] else ""
                    if any(word in first_cell for word in ["total", "suma", "subtotal"]):
                        numeric_cells = sum(1 for cell in row[1:] if self._is_numeric(cell))
                        if numeric_cells >= 2:  # Al menos 2 celdas numéricas después de "Total"
                            is_footer = True

                if is_footer:
                    data_end_row = row_num - 1
                    break
                else:
                    last_data_row = row_num

        return min(data_end_row, last_data_row)

    def _detect_excel_format(self, headers):
        """Detectar qué formato de Excel estamos usando"""
        # Filtrar headers válidos (no nulos, no vacíos, no placeholders)
        headers_valid = [
            h.lower().strip()
            for h in headers
            if h
            and str(h).strip()
            and not str(h).startswith("Empty_Column_")
            and not str(h).startswith("Column_")
            and str(h).strip().lower() != "none"
        ]

        # Formato 3: Petromayab (específico) - debe tener acentos y estructura simple
        format3_indicators = ["vehículo", "vehiculo", "placas", "odómetro", "odometro", "no economico"]
        format3_score = sum(1 for indicator in format3_indicators for header in headers_valid if indicator in header)

        # Formato Cargas: Cargas Odómetros - indicadores exactos del archivo
        formatcargas_indicators = [
            "nota_id",
            "Placas",
            "vehiculo",
            "Odometro",
            "Tipo Mov",
            "id_cliente",
            "razon_social",
            "flotilla_id",
            "desc flotilla",
            "estacion",
        ]
        formatcargas_score = sum(
            1 for indicator in formatcargas_indicators for header in headers_valid if indicator in header
        )

        # Indicadores específicos exactos
        has_vehiculo_acento = any("vehículo" in header for header in headers_valid)
        has_no_economico = any("no economico" in header for header in headers_valid)

        # Para cargas - buscar indicadores exactos del archivo analizado
        has_nota_id = any("nota_id" in header for header in headers_valid)
        has_tipo_mov = any("Tipo Mov" in header for header in headers_valid)
        has_id_cliente = any("Id_Cliente" in header for header in headers_valid)
        has_razon_social = any("razon_social" in header for header in headers_valid)
        has_flotilla_id = any("flotilla_id" in header for header in headers_valid)

        # Lógica de detección mejorada - más específica para cargas
        if has_vehiculo_acento and has_no_economico:
            return "formato_petromayab"
        elif has_nota_id and has_tipo_mov:
            return "formato_petrobol"
        elif has_nota_id and (has_id_cliente or has_razon_social or has_flotilla_id):
            return "formato_petrobol"
        elif formatcargas_score >= 4:
            return "formato_petrobol"
        elif format3_score >= 3:
            return "formato_petromayab"
        else:
            return "formato_generico"

    def _get_column_mappings(self, excel_format):
        """Obtener mapeo de columnas según el formato detectado"""

        if excel_format == "formato_petrobol":
            return {
                # Formato 1 - Original
                "nota_id": "nota_id",
                "Nota_id": "nota_id",
                "Placas": "vehicle",
                "placas": "vehicle",
                "descripcion": "description",
                "producto": "description",
                "Fecha": "date",
                "fecha": "date",
                "Cantidad": "quantity",
                "cantidad": "quantity",
                "Monto": "amount",
                "monto": "amount",
                "Odometro": "odometer",
                "odometro": "odometer",
                "Odómetro": "odometer",
                "Rendimiento": "performance",
                "odo inicial": "initial_odometer",
                "odo_inicial": "initial_odometer",
                "inicial": "initial_odometer",
            }

        elif excel_format == "formato_petromayab":
            return {
                # Formato 2 - Alternativo
                "Tarjeta": "nota_id",
                "tarjeta": "nota_id",
                "Placas": "vehicle",
                "Placa": "vehicle",
                "placas": "vehicle",
                "desc": "description",
                "Descripcion": "description",
                "description": "description",
                "producto": "description",
                "Fecha": "date",
                "fecha": "date",
                "combustible": "product",
                "Producto": "product",
                "product": "product",
                "Cantidad": "quantity",
                "cantidad": "quantity",
                "Monto": "amount",
                "monto": "amount",
                "Odómetro": "odometer",
                "odometro": "odometer",
                "odómetro": "odometer",
                "recorrido": "performance",
                "Rendimiento": "performance",
            }
        # Mapeo por defecto
        return {
            "nota_id": "nota_id",
            "placas": "vehicle",
            "placa": "vehicle",
            "descripcion": "description",
            "fecha": "date",
            "producto": "product",
            "cantidad": "quantity",
            "monto": "amount",
            "Odometro": "odometer",
            "rendimiento": "performance",
            "inicial": "initial_odometer",
        }

    def _read_xlsx_file(self, file_content):
        """Leer archivo Excel y retornar datos estructurados con detección automática"""
        workbook = openpyxl.load_workbook(io.BytesIO(file_content))
        sheet = workbook.active

        # Detectar fila de encabezados
        header_row = self._detect_header_row(sheet)

        # Detectar dónde termina la data (antes del footer)
        data_end_row = self._detect_footer_start(sheet, header_row)

        headers = []
        rows = []

        for row_num, row in enumerate(sheet.iter_rows(values_only=True), 1):
            # Procesar encabezados
            if row_num == header_row:
                # Para columnas fusionadas, mantener todos los headers pero marcar los vacíos
                headers = []
                for i, cell in enumerate(row):
                    if cell is not None and str(cell).strip():
                        headers.append(str(cell).strip())
                    else:
                        headers.append(f"Empty_Column_{i+1}")  # Placeholder más claro
                continue

            # Saltar filas antes del encabezado
            if row_num < header_row:
                continue

            # Parar antes del footer
            if row_num > data_end_row:
                break

            # Saltar filas vacías
            if not any(row):
                continue

            # Convertir valores de celdas a strings legibles
            row_data = []
            for cell in row:
                if cell is None:
                    row_data.append("")
                elif isinstance(cell, datetime):
                    row_data.append(cell.strftime("%Y-%m-%d"))
                else:
                    row_data.append(str(cell))

            rows.append(row_data)

        return {
            "headers": headers,
            "rows": rows,
            "header_row": header_row,
            "data_end_row": data_end_row,
            "footer_excluded": sheet.max_row - data_end_row,
        }

    def _process_xlsx_file(self, file_content):
        """Procesar archivo Excel con detección automática de formato"""
        workbook = openpyxl.load_workbook(io.BytesIO(file_content))
        sheet = workbook.active

        # Detectar fila de encabezados
        header_row = self._detect_header_row(sheet)

        # Detectar dónde termina la data (antes del footer)
        data_end_row = self._detect_footer_start(sheet, header_row)

        lines_data = []
        headers = None

        for row_num, row in enumerate(sheet.iter_rows(values_only=True), 1):
            # Procesar encabezados
            if row_num == header_row:
                headers = [str(cell).strip().lower() if cell else "" for cell in row]
                continue

            # Saltar filas antes del encabezado
            if row_num < header_row:
                continue

            # Parar antes del footer
            if row_num > data_end_row:
                break

            # Saltar filas vacías
            if not any(row):
                continue

            line_data = self._parse_row_data(headers, row, row_num)
            if line_data:
                lines_data.append(line_data)

        return lines_data

    def _parse_row_data(self, headers, row, row_num):
        """Parsear datos de una fila específicamente para las columnas del Excel"""
        try:
            # Mapear columnas específicas de tu Excel
            # column_map = {
            #    "nota_id": "nota_id",
            #    "placas": "vehicle",
            #    "placa": "vehicle",
            #    "descripcion": "description",
            #    "fecha": "date",
            #    "producto": "product",
            #    "cantidad": "quantity",
            #    "monto": "amount",
            #    "odometro": "odometer",
            #    "rendimiento": "performance",
            #    "odo inicial": "initial_odometer",
            #    "odo_inicial": "initial_odometer",
            #    "inicial": "initial_odometer",
            # }
            # Detectar formato y obtener mapeo
            excel_format = self._detect_excel_format(headers)
            column_map = self._get_column_mappings(excel_format)

            data = {}
            odometer_found = False

            for i, header in enumerate(headers):
                if i >= len(row):
                    continue

                header_clean = header.lower().strip()
                value = row[i]
                if header_clean.startswith("empty_column_") or not header_clean:
                    continue
                if header_clean in column_map:
                    field = column_map[header_clean]

                    if field == "nota_id":
                        data["nota_id"] = str(value).strip() if value else ""
                    elif field == "vehicle":
                        plate_cleaned = self._clean_license_plate(value)
                        data["vehicle"] = plate_cleaned
                    elif field == "description":
                        data["description"] = str(value).strip() if value else ""
                    elif field == "date":
                        if isinstance(value, datetime):
                            data["date"] = value.date()
                        elif isinstance(value, str) and value.strip():
                            # Intentar parsear fecha string
                            try:
                                data["date"] = datetime.strptime(value.strip(), "%Y-%m-%d").date()
                            except:
                                try:
                                    data["date"] = datetime.strptime(value.strip(), "%d/%m/%Y").date()
                                except:
                                    try:
                                        data["date"] = datetime.strptime(value.strip(), "%m/%d/%Y").date()
                                    except:
                                        data["date"] = fields.Date.today()
                        else:
                            data["date"] = fields.Date.today()
                    # elif field == "product":
                    #    data["product"] = str(value).strip() if value else ""
                    elif field == "quantity":
                        try:
                            data["quantity"] = float(value) if value else 0.0
                        except:
                            data["quantity"] = 0.0
                    elif field == "amount":
                        try:
                            data["amount"] = float(value) if value else 0.0
                        except:
                            data["amount"] = 0.0
                    elif field == "odometer":
                        try:
                            data["odometer"] = self._parse_float_value(value)
                        except:
                            data["odometer"] = 0.0
                    elif field == "initial_odometer":
                        try:
                            data["initial_odometer"] = float(value) if value else 0.0
                        except:
                            data["initial_odometer"] = 0.0
            # raise ValidationError(_("Datos leidos " + str(data)))
            return data if data else None

        except Exception as e:
            raise ValidationError(_("Error parsing row %d: %s") % (row_num, str(e)))

    def _clean_license_plate(self, plate):
        """Limpiar y normalizar la placa del vehículo"""
        if not plate:
            return ""
        return plate.strip().upper().replace("-", "").replace(" ", "")

    def _search_vehicle_by_plate(self, plate):
        """Buscar vehículo por placa limpia"""
        cleaned_plate = self._clean_license_plate(plate)
        vehicle = self.env["fleet.vehicle"].search([("license_plate", "=", cleaned_plate)], limit=1)
        if not vehicle:
            raise ValidationError(_("Vehiculo con placas '%s' no identificadas.") % cleaned_plate)
        return vehicle

    def _create_odometer_line(self, line_data):
        """Crear línea de odómetro con los nuevos campos"""

        if not line_data.get("vehicle"):
            raise ValidationError(_("Error al obtener vehiculo del excel%s") % line_data.get("vehicle"))
        # Buscar vehículo por placa
        vehicle = self.env["fleet.vehicle"].search([("license_plate", "=", line_data.get("vehicle"))], limit=1)

        if not vehicle:
            # Intentar búsqueda por nombre si no se encuentra por placa
            vehicle = self.env["fleet.vehicle"].search([("name", "ilike", line_data.get("vehicle"))], limit=1)

        if not vehicle:
            raise ValidationError(_("Vehiculo no localizado con la placa: %s") % line_data.get("vehicle"))

        # Validar odómetro
        if line_data.get("odometer", 0) <= 0:
            raise ValidationError(_("Invalido odometro: %s") % line_data.get("odometer"))

        # El n_ticket_partner será la nota_id si existe, sino vacío
        ticket_value = line_data.get("nota_id", "")
        # Crear línea con todos los campos
        line_id = self.env["fleet.batch.odometer.line"].create(
            {
                "batch_id": self.id,
                "vehicle_id": vehicle.id,
                "odometer_value": line_data.get("odometer"),
                "date": line_data.get("date"),
                "n_ticket_partner": ticket_value,
                "total_amount": line_data.get("amount", 0.0),
                # Nuevos campos del Excel
                # "nota_id": line_data.get("nota_id", ""),
                "description": line_data.get("description", ""),
                # "product": line_data.get("product", ""),
                "quantity": line_data.get("quantity", 0.0),
                # "performance": line_data.get("performance", 0.0),
                "initial_odometer": line_data.get("initial_odometer", 0.0),
            }
        )

    def _create_vehicle_odometers(self):
        """Crear registros de odómetro en fleet.vehicle.odometer"""

        created_odometer_ids = []
        for line in self.line_ids:
            try:
                odometer = self.env["fleet.vehicle.odometer"].create(
                    {
                        "vehicle_id": line.vehicle_id.id,
                        "value": line.odometer_value,
                        "date": line.date,
                        "driver_id": line.vehicle_id.driver_id.id if line.vehicle_id.driver_id else False,
                        "batch_id": self.id,
                    }
                )
                created_odometer_ids.append(odometer.id)
            except Exception as e:
                _logger.error("Error al crear el odometro para el vehiculo %s: %s", line.vehicle_id.name, str(e))
                raise ValidationError(
                    _("Error al crear el odometro para el vehiculo %s: %s") % (line.vehicle_id.name, str(e))
                )
        if created_odometer_ids:
            self.odometers_ids = [(6, 0, created_odometer_ids)]
            self.message_post(body=_("Se crearon %d registros de odómetro correctamente.") % len(created_odometer_ids))
            _logger.info("Creados %d odómetros para el lote %s", len(created_odometer_ids), self.name)

    def action_view_odometers(self):
        """Acción para ver los odómetros creados desde este batch"""
        action = self.env["ir.actions.actions"]._for_xml_id("fleet.fleet_vehicle_odometer_action")
        action["domain"] = [("batch_id", "=", self.id)]
        action["context"] = {"create": False, "edit": False}
        return action

    def _parse_float_value(self, value):
        """Parsear valores numéricos de manera flexible"""
        if value is None or value == "":
            return 0.0

        try:
            # Limpiar el valor (quitar comas, espacios, etc.)
            if isinstance(value, str):
                value = value.replace(",", "").replace(" ", "").strip()
            return float(value) if value else 0.0
        except:
            return 0.0


class BatchOdometerLine(models.Model):
    _name = "fleet.batch.odometer.line"
    _description = "Fleet Batch Odometer Line"

    batch_id = fields.Many2one("fleet.batch.odometer", string="Batch", required=True, ondelete="cascade")
    vehicle_id = fields.Many2one("fleet.vehicle", string="Vehiculo", required=True)
    odometer_value = fields.Float("Odometro", required=True)
    description = fields.Char("Descripción")
    quantity = fields.Float(
        "Cantidad L",
        default=0.0,
    )
    date = fields.Date("Fecha", required=True)
    n_ticket_partner = fields.Char(string="No. Ticket", required=True)
    total_amount = fields.Float("Total", required=True)
    initial_odometer = fields.Float("Odómetro Inicial", default=0.0)
    performance = fields.Float("Rendimiento km/l", default=0.0, compute="_compute_performance", store=True)
    km_traveled = fields.Float("Km Recorridos", compute="_compute_km_traveled", store=True)
    currency_id = fields.Many2one("res.currency", string="Moneda", related="batch_id.currency_id", store=True)

    @api.depends("odometer_value", "initial_odometer")
    def _compute_km_traveled(self):
        for line in self:
            if line.initial_odometer > 0:
                line.km_traveled = line.odometer_value - line.initial_odometer
            else:
                line.km_traveled = 0.0

    @api.depends("odometer_value", "initial_odometer", "quantity")
    def _compute_performance(self):
        for line in self:
            if line.initial_odometer > 0:
                line.performance = (line.odometer_value - line.initial_odometer) / line.quantity
            else:
                line.performance = 0.0

    @api.constrains("odometer_value")
    def _check_odometer_value(self):
        for line in self:
            if line.odometer_value < 0:
                raise ValidationError(_("Odometro no puede ser negativo."))

            # Verificar que el odómetro sea mayor al inicial si existe
            # if line.initial_odometer > 0 and line.odometer_value <= line.initial_odometer:
            #    raise ValidationError(
            #        _("Final odometer (%s) must be greater than initial odometer (%s) for vehicle %s.")
            #        % (line.odometer_value, line.initial_odometer, line.vehicle_id.name)
            #   )

            # Verificar que el odómetro sea mayor al último registrado
            last_odometer = self.env["fleet.vehicle.odometer"].search(
                [("vehicle_id", "=", line.vehicle_id.id), ("date", "<=", line.date)],
                order="date desc, id desc",
                limit=1,
            )

            if last_odometer and line.odometer_value <= last_odometer.value:
                raise ValidationError(
                    _("Odometro (%s) debe ser mas grande que el registro anterior (%s) para el sig. vehiculo %s.")
                    % (line.odometer_value, last_odometer.value, line.vehicle_id.name)
                )
